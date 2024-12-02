import os

import openpyxl
import pandas as pd
import shutil
import re
import win32com.client as win32
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from data_processing import main, apply_structure_and_sorting
from testreport import create_od_percent_table

def copy_excel_file(source_file, destination_file):
    """
    Копирует Excel файл с сохранением форматирования, стилей и эффектов.
    """
    try:
        shutil.copy2(source_file, destination_file)
    except Exception as e:
        print(f"Ошибка при копировании файла: {e}")
        raise

def create_columns_with_pywin32(excel_file, sheet_name, start_col, num_columns):
    """
    Создает столбцы в указанном листе Excel файла с использованием pywin32 и метода Columns.Insert.
    """
    xlApp = win32.Dispatch('Excel.Application')
    xlApp.Visible = False
    wb = xlApp.Workbooks.Open(excel_file)
    ws = wb.Sheets(sheet_name)

    # Перемещение к начальной колонке (игнорируя строку)
    ws.Cells(1, start_col).Select()

    for _ in range(num_columns):
        ws.Columns(start_col).Insert(Shift=win32.constants.xlToRight)
        start_col += 1

    wb.Save()
    wb.Close()
    xlApp.Quit()

def create_columns_and_insert_headers_and_data(excel_file, address, df, total_added_columns, initial_headers, is_overall_summary=False):
    """
    Создает столбцы в указанном листе Excel файла с использованием адреса, вставляет заголовки и данные.
    """
    sheet_name, start_row, start_col = address  # Игнорируем строку
    num_columns = len(initial_headers)  # Используем начальный порядок столбцов
    start_col += total_added_columns

    # Создание столбцов с использованием pywin32
    create_columns_with_pywin32(excel_file, sheet_name, start_col, num_columns)

    wb = load_workbook(excel_file)
    ws = wb[sheet_name]

    # Вставка заголовков в порядке initial_headers
    header_row = start_row + 1 if is_overall_summary else start_row
    red_headers = ["Не дозвон", "Бросил трубку", "Другой номер", "Не знаком с клиентом", "Дело в суде", "Клиент умер", "Отказывается от оплаты", "Отказывается от разговора"]

    for i, header in enumerate(initial_headers):
        cell = ws.cell(row=header_row, column=start_col + i)
        cell.value = header

    # Форматирование границ и закрашивание заголовков
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    fill_blue = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for i in range(num_columns):
        cell = ws.cell(row=header_row, column=start_col + i)
        cell.border = thin_border
        if cell.value in red_headers:
            cell.fill = fill_red

    # Объединение ячеек на одну строку выше заголовков
    merge_row = header_row - 1
    start_merge_col = get_column_letter(start_col)
    end_merge_col = get_column_letter(start_col + num_columns - 1)
    ws.merge_cells(f"{start_merge_col}{merge_row}:{end_merge_col}{merge_row}")

    # Закрашивание объединенных ячеек и добавление границ
    for col in range(start_col, start_col + num_columns):
        cell = ws.cell(row=merge_row, column=col)
        cell.fill = fill_blue
        cell.border = thin_border

    # Вставка данных
    for idx, index_value in enumerate(df.index):
        if isinstance(index_value, str) or isinstance(index_value, float):
            cell_address = find_row_in_report(ws, index_value)
            if cell_address:
                col_letter, row_number = re.findall(r'([A-Z]+)([0-9]+)', cell_address)[0]
                row_number = int(row_number)
                for col_idx, header in enumerate(initial_headers):
                    value = df.at[index_value, header] if header in df.columns else None
                    if value is not None:
                        cell = ws.cell(row=row_number, column=start_col + col_idx)
                        cell.value = value
                        cell.number_format = 'General'  # Устанавливаем общий формат
    wb.save(excel_file)
    wb.close()

def find_row_in_report(ws, search_value):
    """
    Находит строку в листе Excel файла, содержащую указанное значение и возвращает адрес ячейки.
    """
    for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row_cells:
            if cell.value == search_value:
                return cell.coordinate
    return None

def main_script(input_file, sheet_name, report_file, report_sheet, output_file):
    try:
        division_column = 'Деления'
        type_column = 'ТИП кредита'
        product_kind_column = 'Вид продукта'
        result_column = 'Результат'
        header_row = 0

        # Получение обработанных данных (X)
        processed_dataframes = main(input_file, sheet_name, division_column, type_column, product_kind_column, result_column, report_file, report_sheet, header_row)

        if not processed_dataframes:
            print("No data was processed from data_processing module.")
            return

        # Сохранение первоначального порядка заголовков для каждого DataFrame
        initial_headers_dict = {df_name: df.columns.tolist() for df_name, df in processed_dataframes.items()}

        # Проверка наличия заголовков в первом столбце
        for df_name, df in processed_dataframes.items():
            if df.columns[0] == "":
                df.columns = [""] + df.columns[1:].tolist()

        # Получение данных из модуля testreport (Y)
        main_merged_cells = ['30-', '30+', '60+', '90+', '180+', '365+']
        addresses_df = create_od_percent_table(report_file, report_sheet, main_merged_cells)

        if addresses_df.empty:
            print("No data was processed from testreport module.")
            return

        # Преобразование адресов вставки в формат словаря
        addresses = {}
        for index, row in addresses_df.iterrows():
            category = row['Category']
            address = row['Address']
            sheet_name = report_sheet  # Все адреса на одном листе
            start_col, start_row = re.findall(r'([A-Z]+)([0-9]+)', address)[0]
            start_col = openpyxl.utils.column_index_from_string(start_col)
            start_row = int(start_row)
            addresses[category] = (sheet_name, start_row, start_col)

        # Копирование исходного файла в новый файл
        copy_excel_file(report_file, output_file)

        total_added_columns = 0

        # Создание столбцов, вставка заголовков и данных для каждой категории
        for category, address in addresses.items():
            if category in processed_dataframes:
                df = processed_dataframes[category]
                initial_headers = initial_headers_dict[category]
                is_overall_summary = category == 'Общий итог'
                create_columns_and_insert_headers_and_data(output_file, address, df, total_added_columns, initial_headers, is_overall_summary)
                total_added_columns += len(initial_headers)  # Учитываем все добавленные столбцы

        print("Процесс завершен успешно.")
    except Exception as e:
        print(f"Произошла ошибка: {e}")

if __name__ == "__main__":
    base_dir = os.path.dirname(os.path.abspath(__file__))

    # Файл с исходными данными
    input_file = os.path.join(base_dir, 'Книга2.xlsx')
    sheet_name = 'Лист1'

    # Файл для отчёта с ключевыми словами
    report_file = os.path.join(base_dir, 'Отчёт 21.06.2024.xlsx')
    report_sheet = 'Сводная погашения NEW'

    # Выходной файл для сохранения обработанных данных
    output_file = os.path.join(base_dir, 'processed_data.xlsx')

    # Основной скрипт для обработки данных и вставки их в скопированный файл
    main_script(input_file, sheet_name, report_file, report_sheet, output_file)
