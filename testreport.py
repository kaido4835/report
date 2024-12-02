import openpyxl
import pandas as pd
import re
import warnings

# Отключаем предупреждения
warnings.filterwarnings("ignore", category=UserWarning, module='openpyxl')


def extract_date_from_filename(file_path):
    print(f"Extracting date from file path: {file_path}")
    match = re.search(r'\d{2}\.\d{2}\.\d{4}', file_path)
    if match:
        date = match.group(0)
        print(f"Extracted date: {date}")
        return date
    print("No date found in file path.")
    return None


def find_address_for_value(workbook, sheet_name, search_value):
    try:
        sheet = workbook[sheet_name]
    except KeyError:
        print(f'Лист "{sheet_name}" не найден в файле.')
        return None

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            if cell.value == search_value:
                return cell.coordinate

    return None


def find_od_percent_address(workbook, sheet_name, main_merged_cells, target_header):
    try:
        sheet = workbook[sheet_name]
    except KeyError:
        print(f'Лист "{sheet_name}" не найден в файле.')
        return []

    results = []

    def increment_column(column, increment=1):
        # Преобразует букву колонки в следующую
        col_num = openpyxl.utils.column_index_from_string(column) + increment
        return openpyxl.utils.get_column_letter(col_num)

    for merged_cell in sheet.merged_cells.ranges:
        cell_value = sheet.cell(merged_cell.min_row, merged_cell.min_col).value
        if cell_value in main_merged_cells:
            for row in range(merged_cell.max_row + 1, sheet.max_row + 1):
                for col in range(merged_cell.min_col, merged_cell.max_col + 1):
                    cell = sheet.cell(row=row, column=col)
                    if cell.value == target_header:
                        for sub_row in range(cell.row + 1, sheet.max_row + 1):
                            for sub_col in range(cell.column,
                                                 cell.column + (merged_cell.max_col - merged_cell.min_col) + 1):
                                sub_cell = sheet.cell(row=sub_row, column=sub_col)
                                if sub_cell.value == '% ОД к просроченному портфелю':
                                    next_col = increment_column(openpyxl.utils.get_column_letter(sub_cell.column))
                                    results.append({
                                        'Category': cell_value,
                                        'Address': f"{next_col}{sub_cell.row}"
                                    })
                                    break
                            else:
                                continue
                            break
                        break

    return results


def create_od_percent_table(file_path, sheet_name, main_merged_cells):
    def increment_column(column, increment=1):
        # Преобразует букву колонки в следующую
        col_num = openpyxl.utils.column_index_from_string(column) + increment
        return openpyxl.utils.get_column_letter(col_num)

    date_from_filename = extract_date_from_filename(file_path)
    if not date_from_filename:
        print('Дата не найдена в имени файла.')
        return pd.DataFrame()

    target_header = f'Просроченная задолженность на {date_from_filename}'
    total_header = f'Кол-во просроченных анкет на {date_from_filename}'

    try:
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        print(f'Файл "{file_path}" не найден.')
        return pd.DataFrame()

    od_percent_results = find_od_percent_address(workbook, sheet_name, main_merged_cells, target_header)
    total_address = find_address_for_value(workbook, sheet_name, total_header)

    if total_address:
        total_col = re.sub(r'[^A-Z]', '', total_address)
        total_row = re.sub(r'[^0-9]', '', total_address)
        new_total_col = increment_column(total_col, increment=3)
        new_total_address = f"{new_total_col}{total_row}"
        od_percent_results.insert(0, {'Category': 'Общий итог', 'Address': new_total_address})

    # Удаление дубликатов из main_merged_cells
    unique_categories = pd.Series(['Общий итог'] + main_merged_cells).unique().tolist()

    df = pd.DataFrame(od_percent_results)

    # Сортировка по порядку в main_merged_cells
    df['Category'] = pd.Categorical(df['Category'], categories=unique_categories, ordered=True)
    df = df.sort_values('Category').reset_index(drop=True)

    return df
